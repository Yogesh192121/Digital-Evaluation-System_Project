from flask import Flask, render_template, request, redirect,send_from_directory, url_for, session, flash, jsonify, send_file
import sqlite3, os
import pandas as pd
import csv, re
import random, string
from functools import wraps
import smtplib
from email.mime.text import MIMEText
from werkzeug.security import generate_password_hash, check_password_hash
import io
from openpyxl import load_workbook
import uuid
from werkzeug.utils import secure_filename
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.lib import colors


app = Flask(__name__)
app.secret_key = "secret123"

# ================= DB CONNECTION =================
def get_db_connection():
    conn = sqlite3.connect("database.db", timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

# ================= LOG SYSTEM =================
def log_activity(conn, user_id, action, details):
    conn.execute("""
        INSERT INTO activity_logs (user_id, action, details)
        VALUES (?, ?, ?)
    """, (user_id, action, details))

# ================= AUTH DECORATORS =================
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Access denied")
            return redirect("/")
        return f(*args, **kwargs)
    return decorated_function


def faculty_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "faculty":
            flash("Access denied")
            return redirect("/")
        return f(*args, **kwargs)
    return decorated_function


def invigilator_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "invigilator":
            flash("Access denied")
            return redirect("/")
        return f(*args, **kwargs)
    return decorated_function

# ================= LOGIN =================

@app.route("/", methods=["GET", "POST"])
def login():

    conn = get_db_connection()

    # ----------------------------
    # BUILD SUBJECT MAPPING
    # ----------------------------
    subjects_by_dept = {}
    departments_set = set()

    courses = conn.execute("SELECT * FROM courses").fetchall()

    for c in courses:
        dept = c["department"]
        departments_set.add(dept)

        if dept not in subjects_by_dept:
            subjects_by_dept[dept] = []

        subjects_by_dept[dept].append({
            "name": c["course_name"],
            "code": c["course_code"]
        })

    # Convert departments to list
    departments = [{"department": d} for d in departments_set]

    # ----------------------------
    # LOGIN LOGIC
    # ----------------------------
    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]
        

        user = conn.execute("""
        SELECT * FROM users 
        WHERE (username=? OR email=?) AND is_approved=1
        """, (username, username)).fetchone()

        if not user:
            conn.close()
            flash("Invalid credentials or not approved yet")
            return redirect("/")

        if not check_password_hash(user["password"], password):
            conn.close()
            flash("Incorrect password")
            return redirect("/")

        # SESSION
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["role"] = user["role"]

        # Force password change
        if user["must_change_password"] == 1:
            return redirect("/change_password")

        # Auto redirect
        if user["role"] == "admin":
            return redirect("/admin_dashboard")

        elif user["role"] == "faculty":
            return redirect("/faculty_dashboard")

        elif user["role"] == "invigilator":
            return redirect("/invigilator_dashboard")
        

        conn.close()
       

    return render_template(
        "login.html",
        subjects_by_dept=subjects_by_dept,
        departments=departments
    )

# ================= ADMIN DASHBOARD =================

@app.route("/admin_dashboard")
@admin_required
def admin_dashboard():

    conn = get_db_connection()

    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    total_courses = conn.execute("SELECT COUNT(*) FROM courses").fetchone()[0]
    total_exams = conn.execute("SELECT COUNT(*) FROM exams").fetchone()[0]
    total_answers = conn.execute("SELECT COUNT(*) FROM student_answers").fetchone()[0]
    evaluated = conn.execute("SELECT COUNT(*) FROM evaluation").fetchone()[0]
    pending = total_answers - evaluated

    assignments = conn.execute("""
SELECT ea.*, 
       c.course_name, 
       e.exam_name,
       f.username AS faculty_name,
       i.username AS invigilator_name

FROM exam_assignments ea
JOIN courses c ON ea.course_id = c.id
JOIN exams e ON ea.exam_id = e.id

LEFT JOIN users f ON ea.assigned_faculty = f.id
LEFT JOIN users i ON ea.assigned_invigilator = i.id
""").fetchall()
    logs = conn.execute("""
        SELECT activity_logs.*, users.username
        FROM activity_logs
        JOIN users ON activity_logs.user_id = users.id
        ORDER BY timestamp DESC
        LIMIT 10
        """).fetchall()

    conn.close()

    return render_template("admin/admin_dashboard.html",
        total_students=total_students,
        total_courses=total_courses,
        total_exams=total_exams,
        total_answers=total_answers,
        evaluated=evaluated,
        pending=pending,
        assignments=assignments,
        logs=logs
        

    )

# ================= FACULTY DASHBOARD =================

@app.route("/faculty_dashboard")
@faculty_required
def faculty_dashboard():

    assignment_id = request.args.get("assignment_id")

    conn = get_db_connection()

    # ✅ GET ALL ASSIGNMENTS (ALWAYS REQUIRED FOR DROPDOWN)
    assignments = conn.execute("""
        SELECT ea.id, ea.year, ea.department,
               c.course_name, e.exam_name
        FROM exam_assignments ea
        JOIN courses c ON ea.course_id = c.id
        JOIN exams e ON ea.exam_id = e.id
        WHERE ea.assigned_faculty = ?
    """, (session.get("user_id"),)).fetchall()

    if assignment_id:
        # ✅ FILTERED DATA

        total = conn.execute(
            "SELECT COUNT(*) FROM student_answers WHERE assignment_id=?",
            (assignment_id,)
        ).fetchone()[0]

        evaluated = conn.execute("""
            SELECT COUNT(*)
            FROM evaluation ev
            JOIN student_answers sa
            ON ev.student_answer_id = sa.id
            WHERE sa.assignment_id=? AND ev.evaluator_id=?
        """, (assignment_id, session.get("user_id"))).fetchone()[0]

    else:
        # ✅ ALL DATA

        total = conn.execute(
            "SELECT COUNT(*) FROM student_answers"
        ).fetchone()[0]

        evaluated = conn.execute(
            "SELECT COUNT(*) FROM evaluation WHERE evaluator_id=?",
            (session.get("user_id"),)
        ).fetchone()[0]

    pending = total - evaluated

    conn.close()

    return render_template(
        "faculty/faculty_dashboard.html",
        total=total,
        evaluated=evaluated,
        pending=pending,
        assignments=assignments,   # ✅ IMPORTANT
        assignment_id=assignment_id
    )
# ================= INVIGILATOR DASHBOARD =================

@app.route("/invigilator_dashboard")
@invigilator_required
def invigilator_dashboard():

    conn = get_db_connection()

    answers = conn.execute("""
    SELECT 
        student_answers.id,
        students.roll_no,
        students.student_name,
        courses.course_code,
        exams.exam_name,
        student_answers.file_path
    FROM student_answers
    JOIN students ON student_answers.student_id = students.id
    JOIN courses ON student_answers.course_id = courses.id
    JOIN exams ON student_answers.exam_id = exams.id
    """).fetchall()

    conn.close()

    return render_template(
        "invigilator/invigilator_dashboard.html",
        answers=answers
    ) 
#=================viwe usres ==================

@app.route("/view_users")
@admin_required
def view_users():


    conn = get_db_connection()

    users = conn.execute("SELECT * FROM users").fetchall()

    conn.close()

    return render_template("admin/view_users.html", users=users)

#================= view students ================

@app.route("/view_students", methods=["GET", "POST"])
def view_students():

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("SELECT DISTINCT department FROM students")
    departments = cursor.fetchall()

    students = []

    if request.method == "POST":

        department = request.form.get("department")
        year = request.form.get("year")

        if year == "all":

            cursor.execute("""
            SELECT * FROM students
            WHERE department=?
            ORDER BY year, roll_no
            """, (department,))

        else:

            cursor.execute("""
            SELECT * FROM students
            WHERE department=? AND year=?
            ORDER BY roll_no
            """, (department, year))

        students = cursor.fetchall()

    conn.close()

    return render_template(
        "admin/view_students.html",
        departments=departments,
        students=students
    )
#===================  edit student ================

@app.route("/edit_student/<int:id>", methods=["GET","POST"])
@admin_required
def edit_student(id):

    conn = get_db_connection()

    if request.method == "POST":

        
        name = request.form["student_name"]
        dept = request.form["department"]
        year = request.form["year"]

        conn.execute("""
        UPDATE students
        SET student_name=?, department=?, year=?
        WHERE id=?
        """, (name, dept, year, id))

        log_activity(conn, session["user_id"], "Edit Student", f"{name} -> {dept} -> {year}")

        flash("Exam created successfully!", "success")

        conn.commit()
        conn.close()

        return redirect("/view_students")

    student = conn.execute(
        "SELECT * FROM students WHERE id=?", (id,)
    ).fetchone()

    conn.close()

    return render_template("admin/edit_student.html", student=student)

#====================== delete student ============================

@app.route("/delete_student/<int:id>")
@admin_required
def delete_student(id):

    conn = get_db_connection()

    conn.execute("DELETE FROM students WHERE id=?", (id,))

    log_activity(conn, session["user_id"], "Delete Student", f"ID {id}")

    conn.commit()
    conn.close()

    return redirect("/view_students")

#================= view exams ===================

@app.route("/view_exams")
@admin_required
def view_exams():

    conn = get_db_connection()

    exams = conn.execute(
        "SELECT * FROM exams"
    ).fetchall()

    conn.close()

    return render_template(
        "admin/view_exams.html",
        exams=exams
    )

#============== view course ===========

@app.route("/manage_courses", methods=["GET", "POST"])
@admin_required
def manage_courses():

    conn = get_db_connection()

    # -------------------------
    # ADD COURSE (POST)
    # -------------------------
    if request.method == "POST":
        course_name = request.form["course_name"].strip()
        course_code = request.form["course_code"].strip()
        dept_form = request.form["department"].strip()

        if not course_name or not course_code or not dept_form:
            flash("All fields are required!")
            return redirect("/manage_courses")

        try:
            conn.execute("""
            INSERT INTO courses (course_name, course_code, department)
            VALUES (?, ?, ?)
            """, (course_name, course_code, dept_form))

            conn.commit()
            flash("Course added successfully!")

        except Exception as e:
            flash("Course code already exists!")

        return redirect("/manage_courses")   # ✅ IMPORTANT

    # -------------------------
    # SEARCH & FILTER (GET)
    # -------------------------
    search = request.args.get("search")
    dept_filter = request.args.get("department")

    query = "SELECT * FROM courses WHERE 1=1"
    params = []

    if search:
        query += " AND (course_name LIKE ? OR course_code LIKE ?)"
        params.extend([f"%{search}%", f"%{search}%"])

    if dept_filter:
        query += " AND department=?"
        params.append(dept_filter)

    courses = conn.execute(query, params).fetchall()

    # -------------------------
    # GET UNIQUE DEPARTMENTS
    # -------------------------
    departments = conn.execute("""
    SELECT DISTINCT department FROM courses
    """).fetchall()

    conn.close()

    return render_template(
        "admin/manage_courses.html",
        courses=courses,
        departments=departments
    )
#=====================================

@app.route("/delete_course/<int:id>")
@admin_required
def delete_course(id):

    conn = get_db_connection()

    conn.execute("DELETE FROM courses WHERE id=?", (id,))
    conn.commit()

    conn.close()

    return redirect("/manage_courses")

#===========================================

@app.route("/edit_course/<int:id>", methods=["GET", "POST"])
@admin_required
def edit_course(id):

    conn = get_db_connection()

    if request.method == "POST":
        conn.execute("""
        UPDATE courses
        SET course_name=?, course_code=?, department=?
        WHERE id=?
        """, (
            request.form["course_name"],
            request.form["course_code"],
            request.form["department"],
            id
        ))

        conn.commit()
        return redirect("/manage_courses")

    course = conn.execute(
        "SELECT * FROM courses WHERE id=?", (id,)
    ).fetchone()

    conn.close()

    return render_template("admin/edit_course.html", course=course)

#=============================================

@app.route("/bulk_upload_courses", methods=["POST"])
@admin_required
def bulk_upload_courses():

    file = request.files["file"]

    if not file:
        flash("No file uploaded")
        return redirect("/manage_courses")

    conn = get_db_connection()
    cursor = conn.cursor()

    csv_file = csv.reader(file.stream.read().decode("utf-8").splitlines())

    next(csv_file)  # skip header

    for row in csv_file:
        course_name, course_code, department = row

        try:
            cursor.execute("""
            INSERT INTO courses (course_name, course_code, department)
            VALUES (?, ?, ?)
            """, (course_name, course_code, department))
        except:
            pass  # skip duplicates

    conn.commit()
    conn.close()

    flash("Courses uploaded successfully!")
    return redirect("/manage_courses")

# ================= CREATE EXAM =================

@app.route("/create_exam", methods=["GET","POST"])
@admin_required
def create_exam():

    conn = get_db_connection()

    departments = conn.execute("SELECT DISTINCT department FROM students").fetchall()
    courses = conn.execute("SELECT * FROM courses").fetchall()
    exams = conn.execute("SELECT * FROM exams").fetchall()

    if request.method == "POST":

        department = request.form["department"]
        year = request.form["year"]
        course_id = request.form["course"]
        exam_id = request.form["exam"]

        conn.execute("""
        INSERT INTO exam_assignments (department, year, course_id, exam_id)
        VALUES (?, ?, ?, ?)
        """, (department, year, course_id, exam_id))
        

        log_activity(conn, session["user_id"], "Create Exam",
                     f"{department} Year {year}")
        
        

        conn.commit()
        conn.close()
        
        return redirect("/admin_dashboard")

    conn.close()
    return render_template("admin/create_exam.html",
                           departments=departments,
                           courses=courses,
                           exams=exams)

# ================= DELETE EXAM =================

@app.route("/delete_exam/<int:id>")
@admin_required
def delete_exam(id):

    conn = get_db_connection()

    # ❗ First delete dependent data (IMPORTANT)
    conn.execute("DELETE FROM student_answers WHERE assignment_id=?", (id,))
    conn.execute("DELETE FROM question_papers WHERE assignment_id=?", (id,))
    conn.execute("DELETE FROM model_answers WHERE assignment_id=?", (id,))
    conn.execute("DELETE FROM evaluation WHERE assignment_id=?", (id,))
    conn.execute("DELETE FROM exam_assignments WHERE id=?", (id,))
    
    log_activity(conn, session["user_id"], "Delete Exam", f"ID {id}")

    conn.commit()
    conn.close()

    flash("🗑️ Exam deleted successfully!", "success")

    return redirect("/admin_dashboard")

# ================= ASSIGN FACULTY =================

@app.route("/assign_faculty/<int:id>", methods=["GET","POST"])
@admin_required
def assign_faculty(id):

    conn = get_db_connection()

    # Get faculty users
    faculty = conn.execute("""
    SELECT * FROM users WHERE role='faculty'
    """).fetchall()

    if request.method == "POST":

        faculty_id = request.form["faculty"]

        conn.execute("""
        UPDATE exam_assignments
        SET assigned_faculty=?, status='assigned'
        WHERE id=?
        """, (faculty_id, id))

        log_activity(conn, session["user_id"],
                 "Assign Faculty",
                 f"Faculty {faculty_id} → Assignment {id}")


        conn.commit()

        return redirect("/admin_dashboard")

    conn.close()

    return render_template(
        "admin/assign_faculty.html",
        faculty=faculty,
        assignment_id=id
    )

#======================= faculty taske ===========================

@app.route("/faculty_tasks")
@faculty_required
def faculty_tasks():

    conn = get_db_connection()

    tasks = conn.execute("""
    SELECT ea.*, c.course_name, e.exam_name
    FROM exam_assignments ea
    JOIN courses c ON ea.course_id = c.id
    JOIN exams e ON ea.exam_id = e.id
    WHERE ea.assigned_faculty = ?
    """, (session["user_id"],)).fetchall()

    conn.close()

    return render_template("faculty/tasks.html", tasks=tasks)

#========================= invigilator exam =================================

@app.route("/invigilator_exams")
@invigilator_required
def invigilator_exams():

    conn = get_db_connection()

    exams = conn.execute("""
    SELECT ea.*, c.course_name, e.exam_name
    FROM exam_assignments ea
    JOIN courses c ON ea.course_id = c.id
    JOIN exams e ON ea.exam_id = e.id
    """).fetchall()

    conn.close()

    return render_template("invigilator/exams.html", exams=exams)

#=====================================================

@app.route("/assign_invigilator/<int:assignment_id>", methods=["GET", "POST"])
@admin_required
def assign_invigilator(assignment_id):

    conn = get_db_connection()

    # ✅ Check assignment exists
    assignment = conn.execute(
        "SELECT * FROM exam_assignments WHERE id = ?",
        (assignment_id,)
    ).fetchone()
    
    if not assignment:
        conn.close()
        flash("Invalid assignment ID", "error")
        return redirect("/admin_dashboard")

    if request.method == "POST":

        invigilator_id = request.form.get("invigilator_id")

        # ✅ Validate input
        if not invigilator_id:
            flash("Please select an invigilator", "error")
            return redirect(f"/assign_invigilator/{assignment_id}")

        # ✅ Verify invigilator exists & is valid
        invigilator = conn.execute("""
            SELECT id FROM users 
            WHERE id = ? 
            AND role = 'invigilator' 
            AND is_approved = 1
        """, (invigilator_id,)).fetchone()

        if not invigilator:
            flash("Invalid invigilator selected", "error")
            return redirect(f"/assign_invigilator/{assignment_id}")

        # ✅ Update assignment
        conn.execute("""
            UPDATE exam_assignments
            SET assigned_invigilator = ?
            WHERE id = ?
        """, (invigilator_id, assignment_id))

        conn.commit()
        conn.close()

        flash("Invigilator assigned successfully", "success")
        return redirect("/admin_dashboard")

    # ✅ Fetch approved invigilators
    invigilators = conn.execute("""
        SELECT id, username 
        FROM users 
        WHERE role = 'invigilator' 
        AND is_approved = 1
        ORDER BY username ASC
    """).fetchall()

    conn.close()

    return render_template(
        "admin/assign_invigilator.html",
        invigilators=invigilators,
        assignment_id=assignment_id
    )

# ================= UPLOAD QUESTION =================

@app.route("/upload_question", methods=["GET", "POST"])
@invigilator_required
def upload_question():

    conn = get_db_connection()
    assignment_id = request.args.get("assignment_id")

    # 🔥 Get assignment (to fetch course_id + exam_id)
    assignment = conn.execute("""
        SELECT * FROM exam_assignments WHERE id=?
    """, (assignment_id,)).fetchone()

    if not assignment:
        conn.close()
        flash("Invalid assignment")
        return redirect("/invigilator_dashboard")

    if request.method == "POST":

        file = request.files.get("file")

        if not file:
            flash("No file selected")
            return redirect(request.url)

        filename = str(uuid.uuid4()) + "_" + secure_filename(file.filename)

        folder = "uploads/question_papers"
        os.makedirs(folder, exist_ok=True)

        full_path = os.path.join(folder, filename)
        file.save(full_path)

        # ✅ STORE CLEAN PATH (NO uploads/)
        db_path = f"question_papers/{filename}"

        conn.execute("""
            INSERT INTO question_papers (course_id, exam_id, file_path, assignment_id)
            VALUES (?, ?, ?, ?)
        """, (
            assignment["course_id"],
            assignment["exam_id"],
            db_path,
            assignment_id,
        ))

        log_activity(conn, session["user_id"],
                     "Upload Question", f"Assignment {assignment_id}")

        conn.commit()
        conn.close()

        flash("Question Paper Uploaded")
        return redirect("/invigilator_dashboard")

    conn.close()

    return render_template(
        "invigilator/upload_question.html",
        assignment_id=assignment_id
    )
# ================= UPLOAD MODEL ANSWER =================

@app.route("/upload_model_answer", methods=["GET", "POST"])
@invigilator_required
def upload_model_answer():

    conn = get_db_connection()
    assignment_id = request.args.get("assignment_id")

    assignment = conn.execute("""
        SELECT * FROM exam_assignments WHERE id=?
    """, (assignment_id,)).fetchone()

    if not assignment:
        conn.close()
        flash("Invalid assignment")
        return redirect("/invigilator_dashboard")

    if request.method == "POST":

        file = request.files.get("file")

        if not file:
            flash("No file selected")
            return redirect(request.url)


        filename = str(uuid.uuid4()) + "_" + secure_filename(file.filename)

        folder = "uploads/model_answers"
        os.makedirs(folder, exist_ok=True)

        full_path = os.path.join(folder, filename)
        file.save(full_path)

        # ✅ CLEAN PATH
        db_path = f"model_answers/{filename}"

        conn.execute("""
            INSERT INTO model_answers (course_id, exam_id, file_path, assignment_id)
            VALUES (?, ?, ?, ?)
        """, (
            assignment["course_id"],
            assignment["exam_id"],
            db_path,
            assignment_id          
        ))

        log_activity(conn, session["user_id"],
                     "Upload Model Answer", f"Assignment {assignment_id}")

        conn.commit()
        conn.close()

        flash("Model Answer Uploaded")
        return redirect("/invigilator_dashboard")

    conn.close()

    return render_template(
        "invigilator/upload_model_answer.html",
        assignment_id=assignment_id
    )

# ================= UPLOAD STUDENT ANSWERS =================

@app.route("/upload_answer", methods=["GET", "POST"])
@invigilator_required
def upload_answer():

    conn = get_db_connection()

    assignment_id = request.args.get("assignment_id")

    # 🔥 Get assignment details
    assignment = conn.execute("""
        SELECT * FROM exam_assignments WHERE id=?
    """, (assignment_id,)).fetchone()

    if not assignment:
        conn.close()
        flash("Invalid assignment")
        return redirect("/invigilator_dashboard")

    # ✅ FILTER STUDENTS (IMPORTANT FIX)
    students = conn.execute("""
        SELECT * FROM students
        WHERE department=? AND year=?
    """, (assignment["department"], assignment["year"])).fetchall()

    # ✅ GET UPLOADED STUDENTS
    answers = conn.execute("""
        SELECT student_id FROM student_answers
        WHERE exam_id=? AND course_id=?
    """, (assignment["exam_id"], assignment["course_id"])).fetchall()

    uploaded_students = [a["student_id"] for a in answers]

    # -------------------------
    # POST (UPLOAD)
    # -------------------------
    if request.method == "POST":

        student_id = request.form["student_id"]
        file = request.files.get("file")

        if not file:
            flash("No file uploaded")
            return redirect(request.url)

        # 🔥 SAVE FILE
        import uuid
        folder = "uploads/student_answers"
        os.makedirs(folder, exist_ok=True)

        filename = str(uuid.uuid4()) + "_" + file.filename
        path = os.path.join(folder, filename)
        file.save(path)

        # 🔥 CHECK EXISTING (PREVENT DUPLICATE)
        existing = conn.execute("""
            SELECT * FROM student_answers
            WHERE student_id=? AND exam_id=? AND course_id=?
        """, (
            student_id,
            assignment["exam_id"],
            assignment["course_id"]
        )).fetchone()

        if existing:
            conn.execute("""
                UPDATE student_answers
                SET file_path=?
                WHERE id=?
            """, (path, existing["id"]))
        else:
            conn.execute("""
                INSERT INTO student_answers
                (student_id, course_id, exam_id, file_path, assignment_id)
                VALUES (?, ?, ?, ?, ?)
            """, (
                student_id,
                assignment["course_id"],
                assignment["exam_id"],
                path,
                assignment_id
            ))

        log_activity(conn, session["user_id"],
                     "Upload Answer",
                     f"Student {student_id} Assignment {assignment_id}")

        conn.commit()
        conn.close()

        flash("Upload successful")
        return redirect(f"/upload_answer?assignment_id={assignment_id}")

    conn.close()

    return render_template(
        "invigilator/upload_answer.html",
        assignment=assignment,
        students=students,
        uploaded_students=uploaded_students,
        assignment_id=assignment_id
    )
#=============== manage exam ===========

@app.route("/manage_exam/<int:id>")
@invigilator_required
def manage_exam(id):

    return render_template("invigilator/manage_exam.html", id=id)

#=============== manage uploads =======

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory('uploads', filename)

#=============== result ===============================

# ================= RESULT DASHBOARD =================

@app.route("/results_dashboard")
@admin_required
def results_dashboard():

    conn = get_db_connection()

    assignments = conn.execute("""
        SELECT ea.id, c.course_name, e.exam_name
        FROM exam_assignments ea
        JOIN courses c ON ea.course_id = c.id
        JOIN exams e ON ea.exam_id = e.id
        ORDER BY ea.id DESC
    """).fetchall()

    conn.close()

    return render_template(
        "admin/results_dashboard.html",
        assignments=assignments
    )


# ================= NUMERICAL REPORT =================

@app.route("/numerical_report/<int:assignment_id>")
@admin_required
def numerical_report(assignment_id):

    conn = get_db_connection()

    total = conn.execute("""
        SELECT COUNT(*) FROM student_answers
        WHERE assignment_id=?
    """, (assignment_id,)).fetchone()[0]

    checked = conn.execute("""
        SELECT COUNT(*) FROM evaluation
        WHERE assignment_id=?
    """, (assignment_id,)).fetchone()[0]

    unchecked = total - checked

    avg = conn.execute("""
        SELECT AVG(total) FROM evaluation
        WHERE assignment_id=?
    """, (assignment_id,)).fetchone()[0]

    top = conn.execute("""
        SELECT s.student_name, e.total, sa.id
        FROM evaluation e
        JOIN student_answers sa ON e.student_answer_id = sa.id
        JOIN students s ON sa.student_id = s.id
        WHERE e.assignment_id=?
        ORDER BY e.total DESC LIMIT 1
    """, (assignment_id,)).fetchone()

    least = conn.execute("""
        SELECT s.student_name, e.total, sa.id
        FROM evaluation e
        JOIN student_answers sa ON e.student_answer_id = sa.id
        JOIN students s ON sa.student_id = s.id
        WHERE e.assignment_id=?
        ORDER BY e.total ASC LIMIT 1
    """, (assignment_id,)).fetchone()

    top5 = conn.execute("""
        SELECT s.student_name, e.total, sa.id
        FROM evaluation e
        JOIN student_answers sa ON e.student_answer_id = sa.id
        JOIN students s ON sa.student_id = s.id
        WHERE e.assignment_id=?
        ORDER BY e.total DESC LIMIT 5
    """, (assignment_id,)).fetchall()

    conn.close()

    return jsonify({
        "total": total,
        "checked": checked,
        "unchecked": unchecked,
        "average": round(avg or 0, 2),
        "top": dict(top) if top else None,
        "least": dict(least) if least else None,
        "top5": [dict(x) for x in top5]
    })


# ================= GRAPHICAL REPORT =================

@app.route("/graphical_report/<int:assignment_id>")
@admin_required
def graphical_report(assignment_id):

    conn = get_db_connection()

    checked = conn.execute("""
        SELECT COUNT(*) FROM evaluation
        WHERE assignment_id=?
    """, (assignment_id,)).fetchone()[0]

    total = conn.execute("""
        SELECT COUNT(*) FROM student_answers
        WHERE assignment_id=?
    """, (assignment_id,)).fetchone()[0]

    conn.close()

    return jsonify({
        "checked": checked,
        "unchecked": total - checked
    })


# ================= MARKS DISTRIBUTION (OUT OF 30) =================

@app.route("/marks_distribution/<int:assignment_id>")
@admin_required
def marks_distribution(assignment_id):

    conn = get_db_connection()

    rows = conn.execute("""
        SELECT total FROM evaluation
        WHERE assignment_id=?
    """, (assignment_id,)).fetchall()

    conn.close()

    ranges = {
        "0-10": 0,
        "11-15": 0,
        "16-20": 0,
        "21-25": 0,
        "26-30": 0
    }

    pass_count = 0
    fail_count = 0

    for r in rows:
        m = r["total"] or 0

        if m <= 10:
            ranges["0-10"] += 1
            fail_count += 1
        elif m <= 15:
            ranges["11-15"] += 1
            pass_count += 1
        elif m <= 20:
            ranges["16-20"] += 1
            pass_count += 1
        elif m <= 25:
            ranges["21-25"] += 1
            pass_count += 1
        else:
            ranges["26-30"] += 1
            pass_count += 1

    return jsonify({
        "ranges": ranges,
        "pass": pass_count,
        "fail": fail_count
    })


# ================= CLICKABLE STUDENT RESULT =================

@app.route("/student_result/<int:answer_id>")
@admin_required
def student_result(answer_id):

    conn = get_db_connection()

    data = conn.execute("""
        SELECT 
            s.student_name,
            s.roll_no,
            e.total,
            e.comments
        FROM evaluation e
        JOIN student_answers sa ON e.student_answer_id = sa.id
        JOIN students s ON sa.student_id = s.id
        WHERE sa.id=?
    """, (answer_id,)).fetchone()

    conn.close()

    return render_template("admin/student_result.html", data=data)


# ================= EXPORT PDF =================

@app.route("/export_pdf")
@admin_required
def export_pdf():

    assignment_id = request.args.get("assignment_id")

    if not assignment_id:
        flash("Invalid request")
        return redirect("/results_dashboard")

    conn = get_db_connection()

    data = conn.execute("""
        SELECT s.student_name, e.total
        FROM evaluation e
        JOIN student_answers sa ON e.student_answer_id = sa.id
        JOIN students s ON sa.student_id = s.id
        WHERE e.assignment_id=?
    """, (assignment_id,)).fetchall()

    conn.close()

    os.makedirs("reports", exist_ok=True)
    file_path = f"reports/report_{assignment_id}.pdf"

    doc = SimpleDocTemplate(file_path)

    table_data = [["Student Name", "Marks"]] + [
        [row["student_name"], row["total"]] for row in data
    ]

    table = Table(table_data)
    table.setStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 1, colors.black)
    ])

    doc.build([table])

    return send_file(file_path, as_attachment=True)


# ================= EXPORT EXCEL =================

@app.route("/export_excel")
@admin_required
def export_excel():

    assignment_id = request.args.get("assignment_id")

    if not assignment_id:
        flash("Invalid request")
        return redirect("/results_dashboard")

    conn = get_db_connection()

    data = conn.execute("""
        SELECT s.student_name, e.total
        FROM evaluation e
        JOIN student_answers sa ON e.student_answer_id = sa.id
        JOIN students s ON sa.student_id = s.id
        WHERE e.assignment_id=?
    """, (assignment_id,)).fetchall()

    conn.close()

    df = pd.DataFrame(data, columns=["Student Name", "Marks"])

    os.makedirs("reports", exist_ok=True)
    file_path = f"reports/report_{assignment_id}.xlsx"

    df.to_excel(file_path, index=False)

    return send_file(file_path, as_attachment=True)

# ================= VIEW ANSWERS =================

@app.route("/view_student_answers")
@faculty_required
def view_answers():

    assignment_id = request.args.get("assignment_id")

    conn = get_db_connection()

    answers = conn.execute("""
    SELECT 
        sa.id,
        s.roll_no,
        s.student_name,
        sa.file_path,

        c.course_name,
        e.exam_name,

        ev.total   -- ✅ USE TOTAL INSTEAD OF marks

    FROM student_answers sa

    JOIN students s ON sa.student_id = s.id
    JOIN exam_assignments ea ON sa.assignment_id = ea.id
    JOIN courses c ON ea.course_id = c.id
    JOIN exams e ON ea.exam_id = e.id

    LEFT JOIN evaluation ev
        ON sa.id = ev.student_answer_id

    WHERE sa.assignment_id = ?
    """, (assignment_id,)).fetchall()

    conn.close()

    return render_template(
        "faculty/view_student_answers.html",
        answers=answers,
        assignment_id=assignment_id
    )
# ================= EVALUATE =================

@app.route("/evaluate/<int:answer_id>", methods=["GET", "POST"])
@faculty_required
def evaluate(answer_id):

    conn = get_db_connection()

    # 🔥 GET ANSWER
    answer = conn.execute("""
        SELECT * FROM student_answers WHERE id=?
    """, (answer_id,)).fetchone()

    if not answer:
        conn.close()
        flash("Answer not found")
        return redirect("/faculty_tasks")

    assignment_id = answer["assignment_id"]

    # ---------------- POST ----------------
    if request.method == "POST":

        def get_mark(field):
            val = request.form.get(field)
            return float(val) if val else 0

        # MARKS
        q1a = get_mark("q1a"); q1b = get_mark("q1b"); q1c = get_mark("q1c")
        q1d = get_mark("q1d"); q1e = get_mark("q1e"); q1f = get_mark("q1f")

        q2a = get_mark("q2a"); q2b = get_mark("q2b"); q2c = get_mark("q2c")

        q3a = get_mark("q3a"); q3b = get_mark("q3b"); q3c = get_mark("q3c")

        total = float(request.form.get("total") or 0)
        comments = request.form.get("comments")

        existing = conn.execute("""
            SELECT * FROM evaluation WHERE student_answer_id=?
        """, (answer_id,)).fetchone()

        if existing:
            conn.execute("""
                UPDATE evaluation SET
                q1a=?, q1b=?, q1c=?, q1d=?, q1e=?, q1f=?,
                q2a=?, q2b=?, q2c=?,
                q3a=?, q3b=?, q3c=?,
                total=?, comments=?, evaluator_id=?
                WHERE student_answer_id=?
            """, (
                q1a,q1b,q1c,q1d,q1e,q1f,
                q2a,q2b,q2c,
                q3a,q3b,q3c,
                total,comments,session["user_id"],
                answer_id
            ))
        else:
            conn.execute("""
                INSERT INTO evaluation (
                    student_answer_id, assignment_id,
                    q1a,q1b,q1c,q1d,q1e,q1f,
                    q2a,q2b,q2c,
                    q3a,q3b,q3c,
                    total,comments,evaluator_id
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                answer_id, assignment_id,
                q1a,q1b,q1c,q1d,q1e,q1f,
                q2a,q2b,q2c,
                q3a,q3b,q3c,
                total,comments,session["user_id"]
            ))

        conn.commit()
        conn.close()

        flash("Evaluation saved")
        return redirect(f"/view_student_answers?assignment_id={assignment_id}")

    # ---------------- GET ----------------

    data = conn.execute("""
        SELECT sa.file_path, s.student_name
        FROM student_answers sa
        JOIN students s ON sa.student_id = s.id
        WHERE sa.id=?
    """, (answer_id,)).fetchone()

    existing = conn.execute("""
        SELECT * FROM evaluation WHERE student_answer_id=?
    """, (answer_id,)).fetchone()

    # 🔥 QUESTION PAPER
    qp = conn.execute("""
        SELECT file_path FROM question_papers
        WHERE assignment_id=?
        ORDER BY id DESC LIMIT 1
    """, (assignment_id,)).fetchone()

    # 🔥 MODEL ANSWER
    ma = conn.execute("""
        SELECT file_path FROM model_answers
        WHERE assignment_id=?
        ORDER BY id DESC LIMIT 1
    """, (assignment_id,)).fetchone()

    conn.close()

    return render_template(
        "faculty/evaluate.html",
        data=data,
        answer=answer,
        existing=existing,
        question_paper=qp["file_path"] if qp else None,
        model_answer=ma["file_path"] if ma else None
    )

#===============bulk upload =============

@app.route("/bulk_upload_students", methods=["POST"])
@admin_required
def bulk_upload_students():

    file = request.files.get("file")

    if not file:
        flash("No file uploaded")
        return redirect("/view_students")

    filename = file.filename.lower()

    conn = get_db_connection()

    try:
        # =========================
        # 📄 CSV FILE SUPPORT
        # =========================
        if filename.endswith(".csv"):

            file_data = file.read()

            try:
                content = file_data.decode("utf-8")
            except UnicodeDecodeError:
                content = file_data.decode("latin-1")

            stream = io.StringIO(content)
            reader = csv.reader(stream)

            next(reader, None)  # skip header

            for row in reader:
                if len(row) < 4:
                    continue

                conn.execute("""
                INSERT INTO students (roll_no, student_name, department, year)
                VALUES (?, ?, ?, ?)
                """, (row[0], row[1], row[2], row[3]))

        # =========================
        # 📊 EXCEL FILE SUPPORT
        # =========================
        elif filename.endswith(".xlsx"):

            wb = load_workbook(file)
            sheet = wb.active

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    continue  # skip header

                if not row or len(row) < 4:
                    continue

                conn.execute("""
                INSERT INTO students (roll_no, student_name, department, year)
                VALUES (?, ?, ?, ?)
                """, (row[0], row[1], row[2], row[3]))

        else:
            flash("Unsupported file format. Upload CSV or Excel (.xlsx)")
            conn.close()
            return redirect("/view_students")

        # =========================
        # ✅ LOGGING
        # =========================
        log_activity(conn, session["user_id"], "Bulk Upload", file.filename)

        conn.commit()
        flash("Students uploaded successfully!")

    except Exception as e:
        conn.rollback()
        flash(f"Upload failed: {str(e)}")

    finally:
        conn.close()

    return redirect("/view_students")

#==================register 

@app.route("/register", methods=["POST"])
def register():

    full_name = request.form["name"]
    email = request.form["email"]
    mobile = request.form["mobile"]
    department = request.form["department"]
    subjects = request.form["subjects"]
    course_codes = request.form.get("course_codes", "")
    address = request.form["address"]
    role = request.form["role"]

    # ----------------------------
    # VALIDATION
    # ----------------------------
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        flash("Invalid email format")
        return redirect("/")

    if not re.match(r"^[0-9]{10}$", mobile):
        flash("Invalid mobile number")
        return redirect("/")

    conn = get_db_connection()

    try:
        if role == "faculty":
            conn.execute("""
            INSERT INTO pending_faculty
            (full_name,email,mobile,department,subjects,course_codes,address)
            VALUES (?,?,?,?,?,?,?)
            """, (full_name, email, mobile, department, subjects, course_codes, address))

        elif role == "invigilator":
            conn.execute("""
            INSERT INTO pending_invigilator
            (full_name,email,mobile,department,address)
            VALUES (?,?,?,?,?)
            """, (full_name, email, mobile, department, address))

        conn.commit()

    except Exception as e:
        print("REGISTER ERROR:", e)
        flash("Something went wrong!")
        conn.close()
        return redirect("/")

    conn.close()

    flash("Registration submitted! Wait for admin approval.")
    return redirect("/")

#========================

def generate_username(full_name, mobile):
    parts = full_name.strip().split()
    first = parts[0][0].lower()
    last = parts[-1].lower()
    return f"{first}{last}{mobile[-2:]}"


def generate_unique_username(full_name, mobile, cursor):
    base = generate_username(full_name, mobile)
    username = base
    count = 1

    while True:
        cursor.execute("SELECT id FROM users WHERE username=?", (username,))
        if not cursor.fetchone():
            return username
        username = f"{base}{count}"
        count += 1


def generate_password():
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(8))

#============================

EMAIL_ADDRESS = "kohalerohit38@gmail.com"
EMAIL_PASSWORD = "bppursgchhmlomky"

def send_email(to, username, password, name):
    subject = "Account Approved"
    body = f"""
Hello {name},

Your account has been approved.

Username: {username}
Password: {password}

Login: http://127.0.0.1:5000
"""

    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = to

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        server.sendmail(EMAIL_ADDRESS, to, msg.as_string())
        server.quit()
    except Exception as e:
        print("Email error:", e)
        
#==============================

@app.route("/pending_requests")
def pending_requests():

    conn = get_db_connection()

    faculty = conn.execute("SELECT * FROM pending_faculty").fetchall()
    invigilator = conn.execute("SELECT * FROM pending_invigilator").fetchall()

    conn.close()

    return render_template(
        "admin/pending_requests.html",
        faculty=faculty,
        invigilator=invigilator
    )

#=================================

@app.route("/approve_faculty/<int:id>")
def approve_faculty(id):

    conn = get_db_connection()
    cursor = conn.cursor()

    f = cursor.execute(
        "SELECT * FROM pending_faculty WHERE id=?", (id,)
    ).fetchone()

    if f:
        username = generate_unique_username(f["full_name"], f["mobile"], cursor)
        password = generate_password()

        # INSERT INTO USERS
        cursor.execute("""
        INSERT INTO users (username,email,password,role,is_approved,must_change_password)
        VALUES (?,?,?,?,?,?)
        """, (
            username,
            f["email"],
            generate_password_hash(password),
            "faculty",
            1,
            1
        ))

        user_id = cursor.lastrowid

        # INSERT INTO PROFILE
        cursor.execute("""
        INSERT INTO faculty_profiles 
        (user_id,full_name,mobile,department,subjects,course_codes,address)
        VALUES (?,?,?,?,?,?,?)
        """, (
            user_id,
            f["full_name"],
            f["mobile"],
            f["department"],
            f["subjects"],
            f["course_codes"],
            f["address"]
        ))

        send_email(f["email"], username, password, f["full_name"])

        cursor.execute("DELETE FROM pending_faculty WHERE id=?", (id,))

        conn.commit()
    flash("Approval Email has been sent to the faculty.")
    conn.close()

    return redirect("/pending_requests")

#====================================

@app.route("/approve_invigilator/<int:id>")
def approve_invigilator(id):

    conn = get_db_connection()
    cursor = conn.cursor()

    i = cursor.execute(
        "SELECT * FROM pending_invigilator WHERE id=?", (id,)
    ).fetchone()

    if i:
        username = generate_unique_username(i["full_name"], i["mobile"], cursor)
        password = generate_password()

        cursor.execute("""
        INSERT INTO users (username,email,password,role,is_approved,must_change_password)
        VALUES (?,?,?,?,?,?)
        """, (
            username,
            i["email"],
            generate_password_hash(password),
            "invigilator",
            1,
            1
        ))

        user_id = cursor.lastrowid

        cursor.execute("""
        INSERT INTO invigilator_profiles
        (user_id,full_name,mobile,department,address)
        VALUES (?,?,?,?,?)
        """, (
            user_id,
            i["full_name"],
            i["mobile"],
            i["department"],
            i["address"]
        ))

        send_email(i["email"], username, password, i["full_name"])

        cursor.execute("DELETE FROM pending_invigilator WHERE id=?", (id,))

        conn.commit()

    conn.close()

    return redirect("/pending_requests")

#===================================

@app.route("/reject_faculty/<int:id>")
@admin_required
def reject_faculty(id):
    conn = get_db_connection()
    conn.execute("DELETE FROM pending_faculty WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect("/pending_requests")

@app.route("/reject_invigilator/<int:id>")
@admin_required
def reject_invigilator(id):
    conn = get_db_connection()
    conn.execute("DELETE FROM pending_invigilator WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect("/pending_requests")

#====================================

@app.route("/change_password", methods=["GET", "POST"])
def change_password():

    if not session.get("user_id"):
        return redirect("/")

    if request.method == "POST":
        old = request.form["old_password"]
        new = request.form["new_password"]
        confirm = request.form["confirm_password"]

        conn = get_db_connection()

        user = conn.execute("""
        SELECT * FROM users WHERE id=?
        """, (session["user_id"],)).fetchone()

        if not check_password_hash(user["password"], old):
            flash("Old password incorrect")
            return redirect("/change_password")

        if new != confirm:
            flash("Passwords do not match")
            return redirect("/change_password")

        conn.execute("""
        UPDATE users
        SET password=?, must_change_password=0
        WHERE id=?
        """, (generate_password_hash(new), session["user_id"]))

        conn.commit()
        conn.close()

        session.clear()
        flash("Password changed. Please login again.")
        return redirect("/")

    return render_template("change_password.html")

import uuid

@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():

    if request.method == "POST":
        email = request.form["email"]

        conn = get_db_connection()

        user = conn.execute("""
        SELECT * FROM users WHERE email=?
        """, (email,)).fetchone()

        if user:
            token = str(uuid.uuid4())

            conn.execute("""
            UPDATE users SET reset_token=? WHERE email=?
            """, (token, email))

            conn.commit()

            reset_link = f"http://127.0.0.1:5000/reset_password/{token}"

            send_email(email, "Password Reset", reset_link)

            flash("Reset link sent to email")

        conn.close()

    return render_template("forgot_password.html")

@app.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_password(token):

    conn = get_db_connection()

    user = conn.execute("""
    SELECT * FROM users WHERE reset_token=?
    """, (token,)).fetchone()

    if not user:
        return "Invalid or expired link"

    if request.method == "POST":
        new_password = request.form["password"]

        conn.execute("""
        UPDATE users SET password=?, reset_token=NULL WHERE id=?
        """, (generate_password_hash(new_password), user["id"]))

        conn.commit()
        conn.close()

        flash("Password reset successful")
        return redirect("/")

    return render_template("reset_password.html")



# ================= LOGOUT =================
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ================= RUN =================
if __name__ == "__main__":
    app.run(debug=True)